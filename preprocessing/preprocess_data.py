# -*- coding: utf-8 -*-
"""Build the single processed input and the isolated silver-label files.

The processed workbook keeps the ten source columns, trims boundary whitespace,
and appends the structured attributes used by the five prediction methods.
Silver labels are written separately so evaluation evidence is never mistaken
for an ordinary model-input column.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
from pathlib import Path

from core.dataset import load
from core.silver import (packcount_invariance_pairs, silver_negatives,
                         silver_pairs)
from core import group as G
from core import norm as N
from paths import (LABELS_DIR, PROCESSED_DATASET_PATH, RAW_DATASET_PATH,
                   ensure_dirs)


DERIVED_COLUMNS = [
    'clean_title', 'ean', 'ean_source', 'brand_canonical', 'brand_key',
    'unit_value', 'unit_uom', 'pack_count', 'product_form', 'form_confidence',
    'origin_normalized', 'variant_signature',
]


def file_sha256(path):
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b''):
            h.update(chunk)
    return h.hexdigest()


def build_processed_records(rows):
    """Return source rows plus deterministic structured preprocessing fields."""
    attrs, _ = G.build_attributes(rows)
    # 🚨 Signatures must use exactly the definition the model uses (G.build_signatures,
    # with attrs carried into the mining key). The old code called this as
    # mine_vocabulary(rows, brand_of) and dropped attrs: the mining key degenerated into
    # the GS1-first brand_id, which fragments a single brand by barcode availability (see
    # the comment on variant.mining_key), so 200/1472 rows ended up with a
    # variant_signature in the xlsx that disagreed with the signature the model used.
    signatures, _, _ = G.build_signatures(rows, attrs)
    out = []
    for row in rows:
        sid = row['sku_id']
        a = attrs[sid]
        rec = dict(row)
        rec.update({
            'clean_title': N.clean_title(row.get('sku_name_chi')),
            'ean': ';'.join(sorted(a['ean'])),
            'ean_source': a['ean_source'] or '',
            'brand_canonical': a['brand_canonical'] or '',
            'brand_key': a['brand_key'] or '',
            'unit_value': a['unit_value'],
            'unit_uom': a['unit_uom'] or '',
            'pack_count': a['pack_count'],
            'product_form': a['form'],
            'form_confidence': a['form_conf'],
            'origin_normalized': '/'.join(sorted(a['origin'])),
            'variant_signature': ';'.join(sorted(signatures[sid])),
        })
        out.append(rec)
    return out, attrs, signatures


def write_processed_workbook(raw_path, output_path, records):
    """Preserve the source workbook and append the preprocessing columns."""
    import openpyxl
    from copy import copy

    wb = openpyxl.load_workbook(raw_path)
    ws = wb[wb.sheetnames[0]]
    source_headers = [c.value for c in ws[1]]
    expected_ids = [str(r['sku_id']) for r in records]
    actual_ids = [str(ws.cell(i, source_headers.index('sku_id') + 1).value).strip()
                  for i in range(2, ws.max_row + 1)]
    if actual_ids != expected_ids:
        raise ValueError('raw workbook row order changed during preprocessing')

    for row_index, rec in enumerate(records, start=2):
        for col_index, header in enumerate(source_headers, start=1):
            value = rec.get(header)
            ws.cell(row_index, col_index).value = value

    start = len(source_headers) + 1
    template = ws.cell(1, len(source_headers))
    for offset, header in enumerate(DERIVED_COLUMNS):
        cell = ws.cell(1, start + offset, header)
        if template.has_style:
            cell._style = copy(template._style)
            cell.font = copy(template.font)
            cell.fill = copy(template.fill)
            cell.border = copy(template.border)
            cell.alignment = copy(template.alignment)
        for row_index, rec in enumerate(records, start=2):
            ws.cell(row_index, start + offset, rec.get(header))

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_silver_labels(rows, attrs, labels_dir):
    """Write positive and negative barcode-derived pairs as separate CSV files."""
    labels_dir = Path(labels_dir)
    labels_dir.mkdir(parents=True, exist_ok=True)
    positives = silver_pairs(rows)
    negatives = silver_negatives(rows, positives, attrs)

    pos_path = labels_dir / 'silver_positive_pairs.csv'
    with pos_path.open('w', newline='', encoding='utf-8') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(['sku_a', 'sku_b', 'label', 'evidence'])
        for a, b in sorted(positives):
            w.writerow([a, b, 1, 'shared_valid_ean13'])

    neg_path = labels_dir / 'silver_negative_pairs.csv'
    with neg_path.open('w', newline='', encoding='utf-8') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(['sku_a', 'sku_b', 'label', 'evidence'])
        for a, b in sorted(negatives):
            w.writerow([a, b, 0, 'disjoint_ean_same_brand_unit_pack'])
    return pos_path, neg_path, len(positives), len(negatives)


def write_pack_pairs(rows, attrs, sigs, path):
    """Freeze the pack-invariance test set into a file. Evaluation only reads it and
    never recomputes it -- otherwise the denominator would drift with the variant
    vocabulary and the metric could improve itself by shrinking the test set."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    pairs = packcount_invariance_pairs(rows, attrs, sigs)
    with path.open('w', newline='', encoding='utf-8') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(['sku_a', 'sku_b', 'label', 'evidence'])
        for a, b in sorted(pairs):
            w.writerow([a, b, 1, 'same_identity_differing_pack_count'])
    return path, len(pairs)


def validate_processed_records(rows, records):
    if len(rows) != len(records):
        raise ValueError('processed row count differs from raw input')
    ids = [r['sku_id'] for r in records]
    if len(ids) != len(set(ids)):
        raise ValueError('processed sku_id values are not unique')
    missing = [sid for sid, rec in zip(ids, records)
               if not rec.get('product_form') or rec.get('pack_count') is None]
    if missing:
        raise ValueError('required preprocessing fields missing: %s' % missing[:5])


def main(raw_path=RAW_DATASET_PATH, output_path=PROCESSED_DATASET_PATH,
         labels_dir=LABELS_DIR):
    ensure_dirs()
    rows = load(raw_path)
    # Variant signatures are built in core/ (G.build_signatures); the xlsx column, the
    # pack-invariance test set and the five prediction methods all share that one
    # definition -- none of the three is allowed to compute its own.
    records, attrs, sigs = build_processed_records(rows)
    validate_processed_records(rows, records)
    write_processed_workbook(raw_path, output_path, records)
    pos_path, neg_path, n_pos, n_neg = write_silver_labels(rows, attrs, labels_dir)
    pack_path, n_pack = write_pack_pairs(rows, attrs, sigs,
                                         Path(labels_dir) / 'pack_invariance_pairs.csv')
    manifest = {
        'raw_path': str(Path(raw_path).resolve()),
        'raw_sha256': file_sha256(raw_path),
        'processed_path': str(Path(output_path).resolve()),
        'processed_sha256': file_sha256(output_path),
        'n_rows': len(records),
        'n_silver_positive_pairs': n_pos,
        'n_silver_negative_pairs': n_neg,
        'n_pack_invariance_pairs': n_pack,
        'note': 'Silver labels are diagnostics, not human-labelled ground truth.',
    }
    manifest_path = Path(labels_dir) / 'preprocessing_manifest.json'
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding='utf-8')
    print('processed:', output_path)
    print('silver labels:', pos_path, neg_path)
    print('pack pairs (frozen):', pack_path, n_pack)
    return manifest


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--raw', default=str(RAW_DATASET_PATH))
    parser.add_argument('--output', default=str(PROCESSED_DATASET_PATH))
    parser.add_argument('--labels-dir', default=str(LABELS_DIR))
    args = parser.parse_args()
    main(args.raw, args.output, args.labels_dir)
