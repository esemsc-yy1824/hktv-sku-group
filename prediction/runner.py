# -*- coding: utf-8 -*-
"""The execution and output contract shared by the five method entry points.

All five methods run the same multi-pass retrieval pipeline; the only difference is how a
listing is represented. The mapping from number to representation is in paths.REPRESENTATION.
"""
from __future__ import annotations

import argparse
import csv
import json
import time
from copy import copy
from pathlib import Path

from paths import (METHOD_LABELS, METHODS, PROCESSED_DATASET_PATH,
                   RAW_DATASET_PATH, method_dir, prediction_workbook)


def validate_group_mapping(path, expected_skus=None):
    with open(path, encoding='utf-8') as f:
        rows = list(csv.DictReader(f))
    required = {'sku_id', 'group_id', 'group_name'}
    if not rows or not required.issubset(rows[0]):
        raise ValueError('%s does not satisfy the prediction schema' % path)
    ids = [r['sku_id'] for r in rows]
    if len(ids) != len(set(ids)):
        raise ValueError('duplicate sku_id in %s' % path)
    if any(not r['group_id'] or not r['group_name'] for r in rows):
        raise ValueError('blank group_id/group_name in %s' % path)
    if expected_skus is not None and set(ids) != set(expected_skus):
        raise ValueError('prediction SKU coverage differs from the raw input')
    return rows


def export_prediction_workbook(method, raw_path=RAW_DATASET_PATH, output_path=None,
                               use_form=True, use_ann=False, use_llm=False):
    """Append exactly group_id/group_name to the ten-column raw workbook."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    output_path = Path(output_path
                       or prediction_workbook(method, use_form, use_ann, use_llm))
    wb = openpyxl.load_workbook(raw_path)
    ws = wb[wb.sheetnames[0]]
    headers = [c.value for c in ws[1]]
    if 'sku_id' not in headers:
        raise ValueError('raw workbook has no sku_id column')
    sku_col = headers.index('sku_id') + 1
    raw_ids = [str(ws.cell(i, sku_col).value).strip() for i in range(2, ws.max_row + 1)]
    prediction_rows = validate_group_mapping(
        method_dir(method, use_form, use_ann, use_llm) / 'groups.csv', raw_ids)
    by_id = {r['sku_id']: r for r in prediction_rows}

    group_id_col = len(headers) + 1
    group_name_col = len(headers) + 2
    template = ws.cell(1, len(headers))
    for col, name in ((group_id_col, 'group_id'), (group_name_col, 'group_name')):
        cell = ws.cell(1, col, name)
        if template.has_style:
            cell._style = copy(template._style)
            cell.font = copy(template.font)
            cell.fill = copy(template.fill)
            cell.border = copy(template.border)
            cell.alignment = copy(template.alignment)
    for row_index, sid in enumerate(raw_ids, start=2):
        ws.cell(row_index, group_id_col, by_id[sid]['group_id'])
        ws.cell(row_index, group_name_col, by_id[sid]['group_name'])
    # Leave the layout of the original ten columns alone; only give the two new result
    # columns enough width to be readable.
    ws.column_dimensions[get_column_letter(group_id_col)].width = 23
    # Disambiguated names carry a distinguishing term; p99 display width is about 76 (CJK
    # characters take two cells), hence 80.
    ws.column_dimensions[get_column_letter(group_name_col)].width = 80
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def run_engine(method, input_path=PROCESSED_DATASET_PATH, offline=False,
               use_form=True, use_ann=False, use_llm=False):
    if not Path(input_path).exists():
        raise FileNotFoundError(
            '%s is missing; run python3 -m preprocessing.preprocess_data first' % input_path)
    from prediction.engine.multipass import main
    main(input_path, method, allow_api=not offline,
         use_form=use_form, use_ann=use_ann, use_llm=use_llm)
    return method_dir(method, use_form, use_ann, use_llm) / 'groups.csv'


def _record_timing(method, export_seconds, total_seconds, reused,
                   use_form=True, use_ann=False, use_llm=False):
    """Add the export time and the overall time to the report.json written by the engine.

    The engine only knows its own segment; "full processing time" runs until the xlsx hits
    disk, and only the runner can see that.
    """
    report_path = method_dir(method, use_form, use_ann, use_llm) / 'report.json'
    if not report_path.exists():
        return
    with report_path.open(encoding='utf-8') as f:
        report = json.load(f)
    timing = report.setdefault('timing', {})
    timing['export_seconds'] = round(export_seconds, 3)
    timing['total_seconds'] = round(total_seconds, 3)
    # When existing debug artefacts are reused the engine never ran, so total is not the
    # cost of a full run
    timing['engine_reused'] = bool(reused)
    with report_path.open('w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)


def run_method(method, input_path=PROCESSED_DATASET_PATH, offline=False,
               reuse_debug=False, use_form=True, use_ann=False, use_llm=False):
    if method not in METHODS:
        raise ValueError('unknown method: %s (choices: %s)' % (method, ', '.join(METHODS)))
    if use_llm and method in ('method_1_lexical', 'method_2_local_tfidf'):
        raise ValueError('--llm-extract applies to the online methods only (3/4/5); '
                         'the whole point of methods 1/2 is that no data leaves the machine')
    started = time.perf_counter()
    debug_groups = method_dir(method, use_form, use_ann, use_llm) / 'groups.csv'
    reused = reuse_debug and debug_groups.exists()
    if not reused:
        run_engine(method, input_path, offline=offline,
                   use_form=use_form, use_ann=use_ann, use_llm=use_llm)
    export_started = time.perf_counter()
    result = export_prediction_workbook(method, use_form=use_form, use_ann=use_ann,
                                        use_llm=use_llm)
    export_seconds = time.perf_counter() - export_started
    total_seconds = time.perf_counter() - started
    _record_timing(method, export_seconds, total_seconds, reused,
                   use_form=use_form, use_ann=use_ann, use_llm=use_llm)
    print('%s -> %s  (%.1fs)' % (METHOD_LABELS[method], result, total_seconds))
    return result


def cli(method):
    parser = argparse.ArgumentParser(description=METHOD_LABELS[method])
    parser.add_argument('--input', default=str(PROCESSED_DATASET_PATH))
    parser.add_argument('--offline', action='store_true')
    parser.add_argument('--reuse-debug', action='store_true',
                        help='reuse an existing debug groups.csv instead of rerunning the model')
    parser.add_argument('--no-form', action='store_true',
                        help='disable product_form; on by default. The impact varies by '
                             'method, see docs/research/ablation_form_ann.txt')
    parser.add_argument('--ann', action='store_true',
                        help='approximate IVF retrieval instead of exact top-k; off by default (1472 rows: same metrics, slower)')
    parser.add_argument('--llm-extract', action='store_true',
                        help='let LLM extraction override the regexes (methods 3/4/5 only, '
                             'artefacts go to output/experiments/); off by default')
    args = parser.parse_args()
    run_method(method, args.input, offline=args.offline, reuse_debug=args.reuse_debug,
               use_form=not args.no_form, use_ann=args.ann, use_llm=args.llm_extract)
