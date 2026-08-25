# -*- coding: utf-8 -*-
"""Repository-level checks for the reorganized layout and checked-in outputs."""
import csv
import json
import unittest
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
METHODS = {
    name: (name, 'soap_sku_data_%s.xlsx' % name)
    for name in ('method_1_lexical', 'method_2_local_tfidf', 'method_3_semantic',
                 'method_4_semantic_image', 'method_5_hybrid_fusion')
}


def workbook_rows(path):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    return rows


class ProjectLayoutTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.raw_rows = workbook_rows(ROOT / 'data/raw/soap_sku_data.xlsx')

    def test_required_layout_exists_and_old_roots_are_gone(self):
        required = [
            'paths.py',
            'prediction/engine/multipass.py',
            'prediction/engine/training.py',
            'prediction/lib/images.py',
            'prediction/lib/ann.py',
            'docs/research/image_signal.py',
            'docs/research/ablation_form_ann.py',
            'docs/research/ablation_form_ann.txt',
            'data/raw/soap_sku_data.xlsx',
            'data/processed/soap_sku_data_preprocessed.xlsx',
            'data/labels/preprocessing_manifest.json',
            'core/dataset.py',
            'core/silver.py',
            'preprocessing/preprocess_data.py',
            'prediction/runner.py',
            'prediction/evaluate_all.py',
            'output/metrics.csv',
            '.github/workflows/ci.yml',
        ]
        for path in required:
            self.assertTrue((ROOT / path).is_file(), path)
        deprecated = [
            'artifacts', 'scripts', 'src', 'tests', 'research', 'evaluation',
            'data/input', 'data/exploration',
            'prediction/hktv_grouping', 'prediction/human_evaluation',
            'prediction/common.py', 'human_evaluation',
            'output/debug/evaluation',
            # Early implementations; must never come back
            'prediction/run_01_phase1_rules.py', 'prediction/run_02_phase2_gbdt.py',
            'prediction/engine/phase1_rules.py', 'prediction/engine/phase2_gbdt.py',
            'prediction/engine/phase2_train.py', 'prediction/engine/method5.py',
            'output/debug/phase1_rules', 'output/debug/phase2_gbdt',
        ]
        for path in deprecated:
            self.assertFalse((ROOT / path).exists(), path)

    def test_preprocessing_does_not_depend_on_prediction(self):
        """Preprocessing may depend only on core/ and paths; a reverse dependency on the
        prediction methods would blur where the labels come from."""
        source = (ROOT / 'preprocessing/preprocess_data.py').read_text(encoding='utf-8')
        for line in source.splitlines():
            # ⚠️ Must strip before testing. The original line.startswith only caught
            # top-level imports; an indented `from prediction... import ...` inside a
            # function body slipped through untouched.
            stripped = line.strip()
            if stripped.startswith(('import ', 'from ')):
                self.assertNotIn('prediction', stripped, line)

    def test_every_method_reports_the_same_metric_keys(self):
        """All five report.json files must expose the same field set, so evaluate_all can
        work from structured data alone."""
        required = {'silver_recall', 'silver_negative_merge_rate',
                    'case_control_precision', 'pack_invariance', 'candidate_recall'}
        for debug_dir, _ in METHODS.values():
            with self.subTest(method=debug_dir):
                path = ROOT / 'output/debug' / debug_dir / 'report.json'
                self.assertTrue(path.is_file(), path)
                with path.open(encoding='utf-8') as handle:
                    report = json.load(handle)
                self.assertIn('metrics', report)
                self.assertTrue(required.issubset(report['metrics']),
                                sorted(required - set(report['metrics'])))

    def test_thresholds_are_selected_by_the_release_gate(self):
        """All five methods use automatic threshold selection, and the selected point
        must genuinely clear the release gate.

        Method 5 used to be pinned at 0.98 (fixed_operating_point) before a business
        decision returned it to automatic selection. This test guards against two
        regressions: the fixed point quietly coming back, or automatic selection picking
        a threshold that does not clear the gate.
        """
        for debug_dir, _ in METHODS.values():
            with self.subTest(method=debug_dir):
                path = ROOT / 'output/debug' / debug_dir / 'report.json'
                with path.open(encoding='utf-8') as handle:
                    report = json.load(handle)
                self.assertEqual(report['threshold_policy'], 'release_gate_auto')
                metrics = report['metrics']
                self.assertLessEqual(metrics['silver_negative_merge_rate'], 0.025)
                self.assertGreaterEqual(metrics['pack_invariance'], 0.98)

    def test_processed_workbook_contract(self):
        rows = workbook_rows(ROOT / 'data/processed/soap_sku_data_preprocessed.xlsx')
        self.assertEqual(len(rows), len(self.raw_rows))
        self.assertEqual(rows[0][:10], self.raw_rows[0])
        self.assertEqual(len(rows[0]), 22)
        eans = [row[11] for row in rows[1:] if row[11] not in (None, '')]
        self.assertTrue(eans)
        self.assertTrue(all(isinstance(ean, str) for ean in eans))

    def test_prediction_workbooks_preserve_raw_and_match_debug_groups(self):
        for method, (debug_dir, output_name) in METHODS.items():
            with self.subTest(method=method):
                rows = workbook_rows(ROOT / 'output' / output_name)
                self.assertEqual(rows[0], self.raw_rows[0] + ('group_id', 'group_name'))
                self.assertEqual([row[:10] for row in rows[1:]], self.raw_rows[1:])
                with (ROOT / 'output/debug' / debug_dir / 'groups.csv').open(
                        encoding='utf-8') as handle:
                    expected = {
                        row['sku_id']: (row['group_id'], row['group_name'])
                        for row in csv.DictReader(handle)
                    }
                actual = {
                    str(row[2]).strip(): (row[10], row[11])
                    for row in rows[1:]
                }
                self.assertEqual(actual, expected)

    def test_metrics_contains_all_five_methods(self):
        with (ROOT / 'output/metrics.csv').open(encoding='utf-8') as handle:
            rows = list(csv.DictReader(handle))
        self.assertEqual(len(rows), 5)
        self.assertEqual({row['output_file'] for row in rows}, {
            output_name for _, output_name in METHODS.values()
        })
        # Runtime and token usage must appear as columns, or costs cannot be compared
        for row in rows:
            self.assertTrue(float(row['total_seconds']) > 0, row['method'])
            self.assertTrue(row['api_tokens'].isdigit(), row['method'])


if __name__ == '__main__':
    unittest.main()
