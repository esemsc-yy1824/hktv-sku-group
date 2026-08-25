# -*- coding: utf-8 -*-
"""SKU workbook reader. Shared entry point for preprocessing and the five prediction methods."""
from __future__ import annotations


def load(path):
    """Read the workbook's first sheet into dicts, trimming whitespace and dropping blank rows."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    rows = []
    for r in it:
        d = {h: (None if v is None else str(v).strip()) for h, v in zip(hdr, r)}
        if any(d.values()):
            rows.append(d)
    wb.close()
    return rows
